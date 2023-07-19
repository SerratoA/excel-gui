import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import datetime
import openpyxl

#Need to add
#Flag for delete
#Multiple row selection for delete and edit
#undo and redo?
#Real time changes 


#Create log file
log_file = "history_log.txt" 
username = ""  # Global variable to store the username
path = "datafull.xlsx"

def convertCSVtoExcel(csv_path, excel_path):
    # Read the CSV file into a pandas DataFrame
    df = pd.read_csv(csv_path)

    # Write the DataFrame to an Excel file
    df.to_excel(excel_path, index=False)

csv_path = "data.csv"
excel_path = "data.xlsx"


def convertExceltoCSV(excel_path, csv_path):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(excel_path)

    # Write the DataFrame to a CSV file
    df.to_csv(csv_path, index=False)

def getUsername():
    global username
    username = simpledialog.askstring("Username", "Please enter your username:")
    if username:
        root.title("Data Management App - User: " + username)
    else:
        root.destroy()


#Load data from Excel Sheet into Treeview
def loadData():
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    treeview.delete(*treeview.get_children())
    
    list_values = list(sheet.values)
    print(list_values)
        
    # Insert the data rows into the Treeview
    for row in list_values[1:]:
        treeview.insert("", "end", values=row)
        
def showSearchResults(results):
    search_window = tk.Toplevel(root)
    search_window.title("Search Results")

    # Create a Treeview in the search window to display the results
    search_treeview = ttk.Treeview(search_window)
    search_treeview.pack()

    # Configure the Treeview columns
    columns = ["Name", "Age", "Subscription", "Employment"]
    search_treeview["columns"] = columns
    search_treeview["show"] = "headings"
    for col in columns:
        search_treeview.heading(col, text=col)
        search_treeview.column(col, width=100)

    # Insert the search results into the Treeview
    for result in results:
        search_treeview.insert("", "end", values=result)

    def copySelectedRow():
        selected_item = search_treeview.focus()
        if selected_item:
            item_values = search_treeview.item(selected_item, "values")
            if item_values:
                name_entry.delete(0, "end")
                name_entry.insert(0, item_values[0])
                age_entry.delete(0, "end")
                age_entry.insert(0, item_values[1])
                status_combobox.set(item_values[2])
                if item_values[3] == "Employed":
                    checkbutton.state(["selected"])
                    a.set(1)
                else:
                    checkbutton.state(["!selected"])
                    a.set(0)


    # Add a "Copy" button
    copy_button = ttk.Button(search_window, text="Copy", command=copySelectedRow)
    copy_button.pack()

def performSearch():
    search_window = tk.Toplevel(root)
    search_window.title("Search")
    
    search_label = ttk.Label(search_window, text="Enter search query:")
    search_label.pack()
    
    search_entry = ttk.Entry(search_window)
    search_entry.pack()
    
    search_button = ttk.Button(search_window, text="Search", command=lambda: searchData(search_entry.get()))
    search_button.pack()

def searchData(x):
    search_text = x  # Get the search text from an entry widget

    # Collect the search results
    results = []
    for item_id in treeview.get_children():
        item_values = treeview.item(item_id)['values']
        if search_text.lower() in [str(value).lower() for value in item_values]:
            results.append(item_values)

    # Show the search results in a pop-up window
    if results:
        showSearchResults(results)
    else:
        messagebox.showinfo("No Results", "No matching results found.")




#Insert row into Excel Sheet and Treeview and clear entry widgets
def insertRow():
    #Get values from entry widgets
    row_values = []
    for key in sorted(entry_widgets.keys()):
        entry_widget = entry_widgets[key]
        value = entry_widget.get()
        row_values.append(value)

    # Get values from checkbuttons in the specified order using the keys
    for key in sorted(checkbutton_widgets.keys()):
        checkbutton = checkbutton_widgets[key]
        value = "Selected" if checkbutton.var.get() else "Not Selected"
        row_values.append(value)

    # Get values from comboboxes in the specified order using the keys
    for key in sorted(combobox_widgets.keys()):
        combobox = combobox_widgets[key]
        value = combobox.get()
        row_values.append(value)

    # Get values from spinboxes in the specified order using the keys
    for key in sorted(spinbox_widgets.keys()):
        spinbox = spinbox_widgets[key]
        value = spinbox.get()
        row_values.append(value)


    # Insert row into Excel Sheet
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.append(row_values)
        workbook.save(path)

    except Exception as e:
        messagebox.showerror("Error", str(e))
        return


    #Insert row into Treeview
    treeview.insert("", "end", values=row_values)

    #Clear entry widgets
    name_entry.delete(0, 'end')
    name_entry.insert(0, 'Name')
    age_entry.delete(0, 'end')
    age_entry.insert(0, 'Age')
    status_combobox.current(0)
    checkbutton.state(['!selected'])

    addHistoryEntry("Inserted row: " + str(row_values))

    # Function to delete selected row from Excel Sheet and Treeview
def deleteRow():
    selected_item = treeview.focus()
    if selected_item:
        confirm = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete this row from the data?")
        if confirm:
            item_values = treeview.item(selected_item, "values")
            if item_values:
            # Delete row from Excel Sheet

                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                row_index = int(treeview.index(selected_item))
                addHistoryEntry("Deleted row: " + str(treeview.item(selected_item)['values']))
                sheet.delete_rows(row_index + 2)  # Adding 2 to compensate for header row and 0-based indexing
                workbook.save(path)

                # Delete row from Treeview
                treeview.delete(selected_item)
                addHistoryEntry("Deleted row: " + str(treeview.item(selected_item)['values']))

    else:
        messagebox.showinfo("No Row Selected", "Please select a row to delete.")


# Edit selected row in Excel Sheet and Treeview
def editRow():
    selected_item = treeview.focus()
    if selected_item:
        item_values = treeview.item(selected_item, "values")
            # Open a new window for editing the row
        edit_window = tk.Toplevel(root)
        edit_window.title("Edit Row")
        row_index = int(treeview.index(selected_item))
            # Create labels and entry widgets for editing the row
        labels = cols
        entries = []
        for i, label in enumerate(labels):
            ttk.Label(edit_window, text=label).grid(row=i, column=0, padx=5, pady=5, sticky="e")
            entry = ttk.Entry(edit_window, width=20)
            entry.insert(0, item_values[i])
            entry.grid(row=i, column=1, padx=5, pady=5)
            entries.append(entry)
            
            # Save button to update the row in Excel Sheet and Treeview
            def saveChanges():
                new_values = [entry.get() for entry in entries]
                try:
                    workbook = openpyxl.load_workbook(path)
                    sheet = workbook.active

                    # Delete the old row
                    sheet.delete_rows(row_index + 2)

                    # Insert the updated row at the same position
                    sheet.insert_rows(row_index + 2)
                    for col_index, value in enumerate(new_values):
                        #print(col_index)
                        sheet.cell(row=row_index + 2, column=col_index+1).value = value

                    workbook.save(path)

                    # Update the row in Treeview
                    treeview.item(selected_item, values=new_values)

                    edit_window.destroy()  # Close the edit window
                    # Add edit entry to history log
                    addHistoryEntry("Edited row: " + str(item_values) + " -> " + str(new_values))
                except Exception as e:
                    print(e)
            
            ttk.Button(edit_window, text="Save", command=saveChanges).grid(row=len(labels), column=0, columnspan=2, padx=5, pady=10)
            
    else:
        messagebox.showinfo("No Row Selected", "Please select a row to edit.")
    
# Function to copy the contents of a row
def copyRow():
    selected_item = treeview.focus()
    if selected_item:
        item_values = treeview.item(selected_item, "values")
        if item_values:
            for key in entry_widgets.keys():
                entry_widget = entry_widgets[key]
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, item_values[key - 1])

            for key in checkbutton_widgets.keys():
                checkbutton_var = checkbutton_widgets[key]
                checkbutton_var.set(item_values[key - 1])

            for key in combobox_widgets.keys():
                combobox_widget = combobox_widgets[key]
                combobox_widget.current(item_values[key - 1])

            for key in spinbox_widgets.keys():
                spinbox_widget = spinbox_widgets[key]
                spinbox_widget.delete(0, tk.END)
                spinbox_widget.insert(0, item_values[key - 1])

        else:
            messagebox.showinfo("No Row Selected", "Please select a row to copy.")
    else:
        messagebox.showinfo("No Row Selected", "Please select a row to copy.")

def addHistoryEntry(entry):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"\n{timestamp} - {username}: {entry}\n"

    with open(log_file, "a") as file:
        file.write(log_entry)
        
# Function to open the log window
#maybe write to a file instead of a text box??
def openLogWindow():
    log_window = tk.Toplevel(root)
    log_window.title("History Logs")
    log_window.geometry("600x400")

    # Create a Text widget in the log window to display the history logs
    log_text = tk.Text(log_window, height=35, width=90)
    log_text.pack()

    # Open the log file and populate the Text widget with its contents
    with open(log_file, "r") as file:
        logs = file.read()
        log_text.insert(tk.END, logs)

def exitApp():
    root.quit()



########---------------------------------------------------------------


#GUI Setup
root = tk.Tk()
root.geometry("1920x1080")
log_window = None
getUsername()


menubar = tk.Menu(root)
root.config(menu=menubar)

gui_menu = tk.Menu(menubar, tearoff=False)
history_menu = tk.Menu(menubar, tearoff=False)

menubar.add_cascade(label="Menu", menu=gui_menu)
menubar.add_cascade(label = "History", menu = history_menu)

# Add search option to the search menu
gui_menu.add_command(label="Search", command=performSearch)
#separator
gui_menu.add_separator()

gui_menu.add_command(label="Exit", command=exitApp)

history_menu.add_command(label="View History", command=openLogWindow)


#Style for Tkinter
root.tk.call('source', 'forest-dark.tcl')
ttk.Style().theme_use('forest-dark')

#Main Frame
frame = ttk.Frame(root, cursor = 'arrow')
frame.pack()

#Widgets on left side of GUI
widgets_entry = ttk.LabelFrame(frame, text='Insert Data Row')
widgets_entry.grid(column=0, row=0, sticky='nsew', padx=10, pady=10)

def create_entry_widget(parent, row, column, width, default_text, key):
    def on_entry_focus_in(event):
        if entry.get() == default_text:
            entry.delete(0, tk.END)

    def on_entry_focus_out(event):
        if entry.get() == "":
            entry.insert(0, default_text)

    entry = ttk.Entry(parent, width=width)
    entry.insert(0, default_text)
    entry.default_text = default_text  # Store the default value as an attribute of the entry widget

    # Bind focus in and focus out events to the Entry widget
    entry.bind('<FocusIn>', on_entry_focus_in)
    entry.bind('<FocusOut>', on_entry_focus_out)

    entry.grid(column=column, row=row, padx=5, pady=5, sticky='ew')
    return entry


# Function to create and configure a checkbutton widget
def create_checkbutton_widget(parent, row, column, text, variable, key):
    checkbutton = ttk.Checkbutton(parent, text=text, variable=variable, onvalue=True, offvalue=False)
    checkbutton.grid(column=column, row=row, padx=5, pady=5, sticky='ew')
    return checkbutton

def create_spinbox_widget(parent, row, column, from_, to, width, default_text, key):
    spinbox = ttk.Spinbox(parent, from_=from_, to=to, width=width)
    spinbox.insert(0, default_text)
    spinbox.default_text = default_text  # Store the default value as an attribute of the spinbox widget

    def on_spinbox_focus_in(event):
        if spinbox.get() == default_text:
            spinbox.delete(0, tk.END)

    def on_spinbox_focus_out(event):
        if spinbox.get() == "":
            spinbox.insert(0, default_text)

    # Bind focus in and focus out events to the Spinbox widget
    spinbox.bind('<FocusIn>', on_spinbox_focus_in)
    spinbox.bind('<FocusOut>', on_spinbox_focus_out)

    spinbox.grid(column=column, row=row, padx=5, pady=5, sticky='ew')
    return spinbox

# Function to create and configure a combobox widget
def create_combobox_widget(parent, row, column, values, width, default_index, key):
    combobox = ttk.Combobox(parent, values=values, state='readonly', width=width)
    combobox.current(default_index)
    combobox.grid(column=column, row=row, padx=5, pady=5, sticky='ew')
    return combobox

# Create entry widgets
custodian_entry = create_entry_widget(widgets_entry, 0, 0, 5, 'Enter Custodian', 1)
source_dest_entry = create_entry_widget(widgets_entry, 1, 0, 5, 'Enter Source Destination', 2)
source_file_entry = create_entry_widget(widgets_entry, 2, 0, 5, 'Enter Source File', 3)
date_format_entry = create_entry_widget(widgets_entry, 3, 0, 5, 'Enter Date Format', 4)
header_delimiter_entry = create_entry_widget(widgets_entry, 4, 0, 5, 'Enter Header Delimiter', 5)
additional_delimiter_entry = create_entry_widget(widgets_entry, 6, 0, 5, 'Enter Additional Header Delimiter', 7)
xls_sheet_name_entry = create_entry_widget(widgets_entry, 9, 0, 5, 'Enter XLS Sheet Name', 10)
zip_file_name_entry = create_entry_widget(widgets_entry, 12, 0, 5, 'Enter Zip File Name', 13)
filter_file_name_entry = create_entry_widget(widgets_entry, 13, 0, 5, 'Enter Filter File Name', 14)
start_string_entry = create_entry_widget(widgets_entry, 1, 1, 5, 'Enter Start String', 19)
first_record_identifier_entry = create_entry_widget(widgets_entry, 5, 1, 5, 'Enter First Record Identifier', 23)
strip_leading_characters_entry = create_entry_widget(widgets_entry, 9, 1, 5, 'Enter Strip Leading Characters', 27)
sequence_entry = create_entry_widget(widgets_entry, 10, 1, 5, 'Enter Sequence', 28)
config_file_entry = create_entry_widget(widgets_entry, 12, 1, 5, 'Enter Config File', 30)
delimiter_entry = create_entry_widget(widgets_entry, 13, 1, 5, 'Enter Delimiter', 31)
filter_value_entry = create_entry_widget(widgets_entry, 0, 2, 5, 'Enter Filter Value', 35)
json_key_name_entry = create_entry_widget(widgets_entry, 2, 2, 5, 'Enter JSON KeyName', 37)
newColumnDate_entry = create_entry_widget(widgets_entry, 4, 2, 5, 'New Column Date', 39)
newColumnIndex_entry = create_entry_widget(widgets_entry, 5, 2, 5, 'New Column Index', 40)
newColumnCount_entry = create_entry_widget(widgets_entry, 6, 2, 5, 'New Column Count', 41)
fileLabel_entry = create_entry_widget(widgets_entry, 9, 2, 5, 'File Label', 44)
server_entry = create_entry_widget(widgets_entry, 10, 2, 5, 'Server', 45)
snowflakeAccount_entry = create_entry_widget(widgets_entry, 12, 2, 5, 'Snowflake Account', 47)
snowflakeAuthenticator_entry = create_entry_widget(widgets_entry, 13, 2, 5, 'Snowflake Authenticator', 48)
snowflakeWarehouse_entry = create_entry_widget(widgets_entry, 14, 2, 5, 'Snowflake Warehouse', 49)
snowflakeDatabase_entry = create_entry_widget(widgets_entry, 15, 2, 5, 'Snowflake Database', 50)
snowflakeSchema_entry = create_entry_widget(widgets_entry, 16, 2, 5, 'Snowflake Schema', 51)
snowflakeFileFormat_entry = create_entry_widget(widgets_entry, 0, 3, 5, 'Snowflake FileFormat', 52)
storedProcedure_entry = create_entry_widget(widgets_entry, 1, 3, 5, 'Stored Procedure', 53)
priority_entry = create_entry_widget(widgets_entry, 3, 3, 5, 'Priority', 55)
notes_entry = create_entry_widget(widgets_entry, 4, 3, 5, 'Notes', 56)

# Create checkbutton widgets
xls_to_csv_var = tk.BooleanVar()
xls_to_csv_checkbutton = create_checkbutton_widget(widgets_entry, 8, 0, 'Convert XLS to CSV', xls_to_csv_var, 9)
unzip_file_var = tk.BooleanVar()
unzip_file_checkbutton = create_checkbutton_widget(widgets_entry, 11, 0, 'Unzip File', unzip_file_var, 12)
combine_files_var = tk.BooleanVar()
combine_files_checkbutton = create_checkbutton_widget(widgets_entry, 14, 0, 'Combine Files', combine_files_var, 15)
remove_header_trailer_var = tk.BooleanVar()
remove_header_trailer_checkbutton = create_checkbutton_widget(widgets_entry, 15, 0, 'Remove Header and Trailer', remove_header_trailer_var, 16)
additional_eol_var = tk.BooleanVar()
additional_eol_checkbutton = create_checkbutton_widget(widgets_entry, 2, 1, 'Additional EOL', additional_eol_var, 20)
remove_additional_eol_var = tk.BooleanVar()
remove_additional_eol_checkbutton = create_checkbutton_widget(widgets_entry, 3, 1, 'Remove Additional EOL', remove_additional_eol_var, 21)
add_record_id_var = tk.BooleanVar()
add_record_id_checkbutton = create_checkbutton_widget(widgets_entry, 4, 1, 'Add Record ID', add_record_id_var, 22)
flatten_file_var = tk.BooleanVar()
flatten_file_checkbutton = create_checkbutton_widget(widgets_entry, 6, 1, 'Flatten File', flatten_file_var, 24)
add_sequence_var = tk.BooleanVar()
add_sequence_checkbutton = create_checkbutton_widget(widgets_entry, 7, 1, 'Add Sequence', add_sequence_var, 25)
fill_with_blank_lines_var = tk.BooleanVar()
fill_with_blank_lines_checkbutton = create_checkbutton_widget(widgets_entry, 8, 1, 'Fill With Blank Lines', fill_with_blank_lines_var, 26)
delimit_fixed_width_var = tk.BooleanVar()
delimit_fixed_width_checkbutton = create_checkbutton_widget(widgets_entry, 11, 1, 'Delimit Fixed Width', delimit_fixed_width_var, 29)
filter_records_var = tk.BooleanVar()
filter_records_checkbutton = create_checkbutton_widget(widgets_entry, 14, 1, 'Filter Records', filter_records_var, 32 )
inverted_filter_var = tk.BooleanVar()
inverted_filter_checkbutton = create_checkbutton_widget(widgets_entry, 15, 1, 'Inverted Filter', inverted_filter_var, 33)
json_scrapping_needed_var = tk.BooleanVar()
json_scrapping_needed_checkbutton = create_checkbutton_widget(widgets_entry, 1, 2, 'JSON Scrapping Needed', json_scrapping_needed_var, 36)
add_column_delimiter_var = tk.BooleanVar()
add_column_delimiter_checkbutton = create_checkbutton_widget(widgets_entry, 3, 2, 'Add Column Delimiter', add_column_delimiter_var, 38)
escapeQuotes_var = tk.BooleanVar()
escapeQuotes_checkbutton = create_checkbutton_widget(widgets_entry, 7, 2, 'Escape Quotes', escapeQuotes_var, 42)
insertFileControl_var = tk.BooleanVar()
insertFileControl_checkbutton = create_checkbutton_widget(widgets_entry, 8, 2, 'Insert File Control', insertFileControl_var, 43)
deleteSourceFile_var = tk.BooleanVar()
deleteSourceFile_checkbutton = create_checkbutton_widget(widgets_entry, 11, 2, 'Delete Source File', deleteSourceFile_var, 46)
flag_entry_var = tk.BooleanVar()
flag_checkbutton = create_checkbutton_widget(widgets_entry, 5, 3, 'Deleteable File', flag_entry_var, 57)

# Create spinbox and combobox widgets
date_position_entry = create_spinbox_widget(widgets_entry, 5, 0, 1, 100, 5, 'Enter Date Position', 6)
num_quarters_entry = create_spinbox_widget(widgets_entry, 10, 0, 1, 100, 5, 'Enter Number of Quarters', 11)
num_header_lines_entry = create_spinbox_widget(widgets_entry, 16, 0, 1, 100, 5, 'Enter Number of Header Lines', 17)
num_trailer_lines_entry = create_spinbox_widget(widgets_entry, 0, 1, 1, 100, 5, 'Enter Number of Trailer Lines', 18)
column_number_entry = create_spinbox_widget(widgets_entry, 16, 1, 1, 100, 5, 'Enter Column Number', 34)

file_type_list = ['Fixed_Width', 'CSV', 'Pipe_Delimited', 'JSON', 'SQL']
file_type_combobox = create_combobox_widget(widgets_entry, 7, 0, file_type_list, 5, 0, 8)
complexity_list = ['Low', 'Medium', 'High']
complexity_combobox = create_combobox_widget(widgets_entry, 2, 3, complexity_list, 5, 0, 54)


#Insert button widget
insert_button = ttk.Button(frame, text='Insert', command = insertRow)
insert_button.grid(column=0, row=1,padx = 5, pady = 5, sticky='ew')

# Create the delete button widget
delete_button = ttk.Button(frame, text='Delete', command=deleteRow)
delete_button.grid(column=0, row=2, padx=5, pady=5, sticky='ew')

# Create the edit button widget
edit_button = ttk.Button(frame, text="Edit", command=editRow)
edit_button.grid(column=0, row=3, padx=5, pady=5, sticky="ew")

# Copy button widget
copy_button = ttk.Button(frame, text="Copy", command=copyRow)
copy_button.grid(column=0, row=4, padx=5, pady=5, sticky="ew")

def clear_all_widgets():
    # Clear entry widgets
    for key in entry_widgets.keys():
        entry_widget = entry_widgets[key]
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, entry_widget.default_text)  # Replace 'Default Text' with the actual default text for each entry

    # Clear checkbutton widgets
    for key in checkbutton_widgets.keys():
        checkbutton_var = checkbutton_widgets[key]
        checkbutton_var.set(False)

    # Clear combobox widgets
    for key in combobox_widgets.keys():
        combobox_widget = combobox_widgets[key]
        combobox_widget.current(0)

    # Clear spinbox widgets
    for key in spinbox_widgets.keys():
        spinbox_widget = spinbox_widgets[key]
        spinbox_widget.delete(0, tk.END)
        spinbox_widget.insert(0, spinbox_widget.default_text)  # Replace 'Default Value' with the actual default value for each spinbox


# Create the "Clear All" button and place it in your GUI
clear_all_button = ttk.Button(frame, text="Clear All", command=clear_all_widgets)
clear_all_button.grid(column=0, row=5, padx=5, pady=5)  # Adjust the row and column values as needed

#Frame for Treeview on right side of GUI
tree_frame = ttk.LabelFrame(frame, text='Data Tree')
tree_frame.grid(column=1, row=0, padx=10, pady=10, sticky='nsew', columnspan=2)

# Scrollbar
tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
tree_scroll_y = ttk.Scrollbar(tree_frame)

tree_scroll_x.pack(side="bottom", fill="x")
tree_scroll_y.pack(side="right", fill="y")



# Treeview
cols = (
    "Custodian",
    "src_Destination_Table",
    "Source_File",
    "Date_Format",
    "Header_Delimiter",
    "Date_Position_OR_Column",
    "Additional_Delimiter",
    "File_Type",
    "XLS_to_CSV",
    "XLS_Sheet_Name",
    "Number_Of_Quarters",
    "Unzip_File",
    "Zip_File_Name",
    "Filter_File_Name",
    "Combine_Files",
    "Remove_Header_Trailer",
    "Num_Header_Lines",
    "Num_Trailer_Lines",
    "Start_String",
    "Additional_EOL",
    "Remove_Additional_EOL",
    "Add_Record_ID",
    "First_Record_Identifier",
    "Flatten_File",
    "Add_Sequence",
    "Fill_With_Blank_Lines",
    "Strip_Leading_Characters",
    "Sequence",
    "Delimit_Fixed_Width",
    "Config_File",
    "Delimiter",
    "Filter_Records",
    "Inverted_Filter",
    "Column_Number",
    "Filter_Value",
    "JSON_Scrapping_Needed",
    "JSON_KeyName",
    "Add_Column_Delimiter",
    "New_Column_Date",
    "New_Column_Index",
    "New_Column_Count",
    "Escape_Quotes",
    "Insert_File_Control",
    "File_Label",
    "Server",
    "Delete_Source_File",
    "Snowflake_Account",
    "Snowflake_Authenticator",
    "Snowflake_Warehouse",
    "Snowflake_Database",
    "Snowflake_Schema",
    "Snowflake_FileFormat",
    "Stored_Procedure",
    "Complexity",
    "Priority",
    "Notes",
)

entry_widgets = {
    1: custodian_entry,
    2: source_dest_entry,
    3: source_file_entry,
    4: date_format_entry,
    5: header_delimiter_entry,
    7: additional_delimiter_entry,
    10: xls_sheet_name_entry,
    13: zip_file_name_entry,
    14: filter_file_name_entry,
    19: start_string_entry,
    23: first_record_identifier_entry,
    27: strip_leading_characters_entry,
    28: sequence_entry,
    30: config_file_entry,
    31: delimiter_entry,
    35: filter_value_entry,
    37: json_key_name_entry,
    39: newColumnDate_entry,
    40: newColumnIndex_entry,
    41: newColumnCount_entry,
    44: fileLabel_entry,
    45: server_entry,
    47: snowflakeAccount_entry,
    48: snowflakeAuthenticator_entry,
    49: snowflakeWarehouse_entry,
    50: snowflakeDatabase_entry,
    51: snowflakeSchema_entry,
    52: snowflakeFileFormat_entry,
    53: storedProcedure_entry,
    55: priority_entry,
    56: notes_entry,
}

checkbutton_widgets = {
    9: xls_to_csv_var,
    12: unzip_file_var,
    15: combine_files_var,
    16: remove_header_trailer_var,
    20: additional_eol_var,
    21: remove_additional_eol_var,
    22: add_record_id_var,
    24: flatten_file_var,
    25: add_sequence_var,
    26: fill_with_blank_lines_var,
    29: delimit_fixed_width_var,
    32: filter_records_var,
    33: inverted_filter_var,
    36: json_scrapping_needed_var,
    38: add_column_delimiter_var,
    42: escapeQuotes_var,
    43: insertFileControl_var,
    46: deleteSourceFile_var,
    54: flag_entry_var,
}

combobox_widgets = {
    8: file_type_combobox,
    54: complexity_combobox,
}

spinbox_widgets = {
    6: date_position_entry,
    11: num_quarters_entry,
    17: num_header_lines_entry,
    18: num_trailer_lines_entry,
    34: column_number_entry,
}

treeview = ttk.Treeview(tree_frame, columns=cols, show="headings", height=40, yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
for col in cols:
    treeview.heading(col, text=col)
    treeview.column(col, width=100, anchor='center')



treeview.pack(expand = True, fill = 'both')


tree_scroll_y.config(command=treeview.yview) 
tree_scroll_x.config(command=treeview.xview)

# Load data into Treeview
loadData()

#Set the weight of rows and columns to make the treeview fit the window
frame.columnconfigure(1, weight=1)
frame.rowconfigure(0, weight=1)


# Run GUI
root.mainloop()
