import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import datetime
import openpyxl
import json

#TODO: 
#Robert: Think of better save methodology, undo feature for the save? undo
#feature for the before save part? 

#Angle, lets try to split the code into at least 5 files 

#Create log file
log_file = "history_log.txt" 
username = ""  # Global variable to store the username
path = "datafull.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

def getUsername():
    global username
    username = simpledialog.askstring("Username", "Please enter your username:")
    if username:
        root.title("Data Management App - User: " + username)
    else:
        root.destroy()


#Load data from Excel Sheet into Treeview
def loadData():
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    treeview.delete(*treeview.get_children())
    
    list_values = list(sheet.values)
#    print(list_values)
        
    # Insert the data rows into the Treeview
    for row in list_values[1:]:
        treeview.insert("", "end", values=row)
        
#Save changes to Excel Sheet and display a message box
def saveChanges():
    workbook.save(path)
    messagebox.showinfo("Success", "Your changes have been saved.")


# Global variable to store the selected item in the search results Treeview
selected_search_item = None

def highlightSelectedRow():
    global selected_search_item_id
    if selected_search_item_id is not None:
        # Find the corresponding row in the main data Treeview and highlight it
        for item_id in treeview.get_children():
            if treeview.item(item_id, "values")[1] == selected_search_item_id:
                treeview.selection_set(item_id)  # Highlight the row
                treeview.see(item_id)  # Scroll to make the selection visible
                break

def showSearchResults(results):
    global selected_search_item_id
    search_window = tk.Toplevel(root)
    search_window.title("Search Results")

    # Create a Treeview in the search window to display the results
    search_treeview = ttk.Treeview(search_window)
    search_treeview.pack()

    # Configure the Treeview columns
    columns = ["Custodian", "Source_Destination_Table", "Source_File", "Date_Format"]
    search_treeview["columns"] = columns
    search_treeview["show"] = "headings"
    for col in columns:
        search_treeview.heading(col, text=col)
        search_treeview.column(col, width=100)

    # Insert the search results into the Treeview
    for result in results:
        search_treeview.insert("", "end", values=result)

    # Add a "Copy" button
    highlight_button = ttk.Button(search_window, text="Highlight", command=highlightSelectedRow)
    highlight_button.pack()

    def on_search_item_selected(event):
        global selected_search_item_id
        selected_item = search_treeview.selection()
        if selected_item:
            selected_search_item_id = search_treeview.item(selected_item, "values")[1]
        else:
            selected_search_item_id = None

    search_treeview.bind("<<TreeviewSelect>>", on_search_item_selected)





def performSearch():
    search_window = tk.Toplevel(root)
    search_window.title("Search")
    
    search_label = ttk.Label(search_window, text="Enter search query:")
    search_label.pack()
    
    search_entry = ttk.Entry(search_window)
    search_entry.pack()
    
    search_button = ttk.Button(search_window, text="Search", command=lambda: searchData(search_entry.get()))
    search_button.pack()

def searchData(x):
    search_text = x  # Get the search text from an entry widget

    # Collect the search results
    results = []
    for item_id in treeview.get_children():
        item_values = treeview.item(item_id)['values']
        if search_text.lower() in [str(value).lower() for value in item_values]:
            results.append(item_values)

    # Show the search results in a pop-up window
    if results:
        showSearchResults(results)
    else:
        messagebox.showinfo("No Results", "No matching results found.")




#Insert row into Excel Sheet and Treeview and clear entry widgets
def insertRow():
    #Get values from entry widgets
    row_values = []
    for key in sorted(entry_widgets.keys()):
        entry_widget = entry_widgets[key]
        value = entry_widget.get()
        row_values.append(value)

    # Insert row into Excel Sheet
    try:
        sheet.append(row_values)
    except Exception as e:
        messagebox.showerror("Error", str(e))
        return

    #Insert row into Treeview
    treeview.insert("", "end", values=row_values)
    addHistoryEntry("\nInserted row: " + str(row_values))

    # Function to delete selected row from Excel Sheet and Treeview
def deleteRow():
    selected_item = treeview.focus()
    if selected_item:
        confirm = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete this row from the data?")
        if confirm:
            item_values = treeview.item(selected_item, "values")
            if item_values:
            # Delete row from Excel Sheet

                row_index = int(treeview.index(selected_item))
                addHistoryEntry("\nDeleted row: " + str(treeview.item(selected_item)['values']))
                sheet.delete_rows(row_index + 2)  # Adding 2 to compensate for header row and 0-based indexing   
                # Delete row from Treeview
                treeview.delete(selected_item)
                addHistoryEntry("\nDeleted row: " + str(treeview.item(selected_item)['values']))

    else:
        messagebox.showinfo("No Row Selected", "Please select a row to delete.")

    
# Edit selected row in Excel Sheet and Treeview
def editRow():
    selected_item = treeview.focus()
    if selected_item:
        item_values = treeview.item(selected_item, "values")
        # Open a new window for editing the row
        edit_window = tk.Toplevel(root)
        edit_window.title("Edit Row")
        row_index = int(treeview.index(selected_item))
        
        # Create labels and entry widgets for editing the row
        labels = columns
        entries = []
        
        # Create a canvas with a scrollbar
        canvas = tk.Canvas(edit_window, width=400, height=300)  # Set the size as needed
        scrollbar = ttk.Scrollbar(edit_window, orient="vertical", command=canvas.yview)
        canvas.config(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Create a frame inside the canvas to hold the entry widgets
        edit_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=edit_frame, anchor="nw")

        for i, label in enumerate(labels):
            ttk.Label(edit_frame, text=label).grid(row=i, column=0, padx=5, pady=5, sticky="e")
            entry = ttk.Entry(edit_frame, width=30)  # Adjust the width as needed
            entry.insert(0, item_values[i])
            entry.grid(row=i, column=1, padx=5, pady=5)
            entries.append(entry)

        # Configure the canvas scrolling region and scroll wheel binding
        edit_frame.update_idletasks()  
        bbox = canvas.bbox(tk.ALL)  
        canvas.configure(scrollregion=bbox, width=400, height=300)
        # Bind the MouseWheel to the canvas for vertical scrolling
        canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1 * (event.delta / 120)), "units"))

        
        # Save button to update the row in Excel Sheet and Treeview
        def saveChanges():
                new_values = [entry.get() for entry in entries]
                # Delete the old row
                sheet.delete_rows(row_index + 2)
                # Insert the updated row at the same position
                sheet.insert_rows(row_index + 2)
                for col_index, value in enumerate(new_values):
                    sheet.cell(row=row_index + 2, column=col_index + 1).value = value
                # Update the row in Treeview
                treeview.item(selected_item, values=new_values)
                edit_window.destroy()  # Close the edit window
                # Add edit entry to history log
                addHistoryEntry("\nEdited row: " + str(item_values) + " -> " + str(new_values))

        ttk.Button(edit_window, text="Save", command=saveChanges).pack(padx=5, pady=5)

    else:
        messagebox.showinfo("No Row Selected", "Please select a row to edit.")
    
# Function to copy the contents of a row
def copyRow():
    selected_item = treeview.focus()
    if selected_item:
        item_values = treeview.item(selected_item, "values")
        if item_values:
            for key in entry_widgets.keys():
                entry_widget = entry_widgets[key]
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, item_values[key - 1])
        else:
            messagebox.showinfo("No Row Selected", "Please select a row to copy.")
    else:
        messagebox.showinfo("No Row Selected", "Please select a row to copy.")

def addHistoryEntry(entry):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"\n{timestamp} - {username}: {entry}\n"

    with open(log_file, "a") as file:
        file.write(log_entry)

def clear_all_widgets():
    # Clear entry widgets
    for key in entry_widgets.keys():
        entry_widget = entry_widgets[key]
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, entry_widget.default_text)  

# Function to open the log window
#maybe write to a file instead of a text box??
def openLogWindow():
    log_window = tk.Toplevel(root)
    log_window.title("History Logs")
    log_window.geometry("600x400")

    # Create a Text widget in the log window to display the history logs
    log_text = tk.Text(log_window, height=35, width=90)
    log_text.pack()

    # Open the log file and populate the Text widget with its contents
    with open(log_file, "r") as file:
        logs = file.read()
        log_text.insert(tk.END, logs)

def aboutPage():
    about_window = tk.Toplevel(root)
    about_window.title("About")
    about_window.geometry("300x200")
    
    about_label = ttk.Label(about_window, text="Data Management Application", font=("Helvetica", 16))
    about_label.pack(pady=20)

    about_text = ttk.Label(about_window, text="Created By:\n Angel Serrato and Chaeil Yun\n\n\nVersion 0.8.0")
    about_text.pack()

    ok_button = ttk.Button(about_window, text="Close", command=about_window.destroy)
    ok_button.pack(pady=10)

def exitApp():
    root.quit()



########---------------------------------------------------------------


#GUI Setup
root = tk.Tk()
root.geometry("1620x980")
log_window = None
getUsername()


menubar = tk.Menu(root)
root.config(menu=menubar)

gui_menu = tk.Menu(menubar, tearoff=False)
history_menu = tk.Menu(menubar, tearoff=False)

menubar.add_cascade(label="Menu", menu=gui_menu)
# menubar.add_cascade(label = "History", menu = history_menu)

# Add abouts section
gui_menu.add_command(label="About", command=aboutPage)

# Add search option to the search menu
gui_menu.add_command(label="Search", command=performSearch)

# history menu 
gui_menu.add_command(label="View History", command=openLogWindow)
#separator
gui_menu.add_separator()

gui_menu.add_command(label="Exit", command=exitApp)


#Style for Tkinter
root.tk.call('source', 'forest-dark.tcl')
ttk.Style().theme_use('forest-dark')

#Main Frame
frame = ttk.Frame(root, cursor = 'arrow')
frame.pack()

widgets_entry_canvas = tk.Canvas(frame, highlightthickness=0)
widgets_entry_canvas.grid(column=0, row=0, sticky='nsew')

#Widgets labelframe inside canvas for the entry widgets with height and width of the canvas
widgets_entry = ttk.LabelFrame(widgets_entry_canvas, text='Insert Data Row')
widgets_entry_canvas.create_window((0, 0), window=widgets_entry, anchor='nw')

# Create vertical scrollbar for the canvas and attach it to the canvas 
entry_scrollbar = ttk.Scrollbar(widgets_entry_canvas, orient='vertical', command=widgets_entry_canvas.yview)
entry_scrollbar.pack(side='right', fill='y')
widgets_entry_canvas.configure(yscrollcommand=entry_scrollbar.set)


# Configure the canvas
widgets_entry_canvas.configure(yscrollcommand=entry_scrollbar.set)
widgets_entry_canvas.bind('<Configure>', lambda e: widgets_entry_canvas.configure(scrollregion=widgets_entry_canvas.bbox('all')))
widgets_entry_canvas.bind_all('<MouseWheel>', lambda e: widgets_entry_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))




def create_entry_widget(parent, row, column, width, default_text):
    def on_entry_focus_in(event):
        if entry.get() == default_text:
            entry.delete(0, tk.END)

    def on_entry_focus_out(event):
        if entry.get() == "":
            entry.insert(0, default_text)

    entry = ttk.Entry(parent, width=width)
    entry.insert(0, default_text)
    entry.default_text = default_text  # Store the default value as an attribute of the entry widget

    # Bind focus in and focus out events to the Entry widget
    entry.bind('<FocusIn>', on_entry_focus_in)
    entry.bind('<FocusOut>', on_entry_focus_out)

    entry.grid(column=column, row=row, padx=5, pady=5, sticky='ew')
    return entry

with open('config.json', 'r') as config_file:
    config_data = json.load(config_file)

columns = config_data['columns']
default_values = config_data['default_values']


entry_widgets = {}
for i, col in enumerate(columns):
    entry_widget = create_entry_widget(widgets_entry, i, 0, 45, default_values.get(col, ''))
    entry_widgets[i+1] = entry_widget
  


#Insert button widget
insert_button = ttk.Button(frame, text='Insert', command = insertRow, width = 45)
insert_button.grid(column=0, row=1,padx = 5, pady = 5)

# Create the delete button widget
delete_button = ttk.Button(frame, text='Delete', command=deleteRow, width = 45)
delete_button.grid(column=0, row=2, padx=5, pady=5)

# Create the edit button widget
edit_button = ttk.Button(frame, text="Edit", command=editRow, width = 45)
edit_button.grid(column=0, row=3, padx=5, pady=5)

# Copy button widget
copy_button = ttk.Button(frame, text="Copy", command=copyRow, width = 45)
copy_button.grid(column=0, row=4, padx=5, pady=5)

save_button = ttk.Button(frame, text="Save", command=saveChanges, width = 45)
save_button.grid(column = 0, row = 6 , padx = 5, pady = 5)


# Create the "Clear All" button and place it in your GUI
clear_all_button = ttk.Button(frame, text="Clear All", command=clear_all_widgets, width = 45)
clear_all_button.grid(column=0, row=5, padx=5, pady=5)  # Adjust the row and column values as needed

#Frame for Treeview on right side of GUI, spans 2 rows
tree_frame = ttk.LabelFrame(frame, text='Data Tree')
tree_frame.grid(column=1, row=0, padx=12, pady=10, sticky='nsew', rowspan=8)

# Scrollbar
tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
tree_scroll_y = ttk.Scrollbar(tree_frame)

tree_scroll_x.pack(side="bottom", fill="x")
tree_scroll_y.pack(side="right", fill="y")

treeview = ttk.Treeview(tree_frame, columns=columns, show="headings", height=40, yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
for col in columns:
    treeview.heading(col, text=col)
    treeview.column(col, width=100, anchor='center')


treeview.pack(expand = True, fill = 'both')


tree_scroll_y.config(command=treeview.yview) 
tree_scroll_x.config(command=treeview.xview)

# Load data into Treeview
loadData()

#Set the weight of rows and columns to make the treeview fit the window
frame.columnconfigure(1, weight=1)
frame.rowconfigure(0, weight=1)

# Run GUI
root.mainloop()
